#!/usr/bin/env python3
"""Extract user grades from team NotationHackathon.xlsx files.

Produces output in the current directory:
- summary_report.xlsx : Excel document with multiple sheets:
  - All Grades
  - All Grades By Domain
  - Top 10 Grades
  - Bottom 10 Grades
  - Team Averages
  - Domain-<DOMAIN> for each domain found
  - Statistics

Assumptions taken from the different Excel files:
- Each team folder contains a file named exactly "NotationHackathon.xlsx"
- Worksheet named "Oral" (case-insensitive) marks the end; all sheets before it are user-grade sheets
- On a user-grade sheet, cell B12 contains the user domain and C19 contains the total grade

This script is defensive and will skip missing files / missing cells while reporting progress.
"""
from pathlib import Path
import sys
import pandas as pd
from openpyxl import load_workbook


def find_team_dirs(root: Path):
    """Return a list of immediate subdirectories of root."""
    return [p for p in sorted(root.iterdir()) if p.is_dir()]


def process_notation_file(path: Path, team_name: str):
    """Process a single NotationHackathon.xlsx file and yield rows.

    Yields dicts with keys: Team Name, User Domain, User Name, User Grade
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "oral":
            # stop processing further sheets
            break
        ws = wb[sheet_name]
        user_name = sheet_name
        domain = ws["B12"].value if ws["B12"] is not None else None
        grade = ws["C19"].value if ws["C19"] is not None else None
        yield {
            "Team Name": team_name,
            "User Domain": domain,
            "User Name": user_name,
            "User Grade": grade,
        }


def extract_all(root: Path):
    rows = []
    missing = []
    not_excel = []
    for team_dir in find_team_dirs(root):
        team_name = team_dir.name
        candidate = team_dir / "NotationHackathon.xlsx"
        if not candidate.exists():
            missing.append(team_name)
            continue
        try:
            for row in process_notation_file(candidate, team_name):
                rows.append(row)
        except Exception as e:
            not_excel.append((team_name, str(e)))
    return rows, missing, not_excel


def main(argv):
    root = Path.cwd()
    print(f"Scanning teams in: {root}")
    rows, missing, errors = extract_all(root)

    print(f"Found {len(rows)} user-grade rows from {len(set(r['Team Name'] for r in rows))} teams.")
    if missing:
        print(f"Warning: {len(missing)} team folders missing NotationHackathon.xlsx: {', '.join(missing)}")
    if errors:
        print("Errors while opening some files:")
        for team, err in errors:
            print(f" - {team}: {err}")

    if not rows:
        print("No rows to write.")
        return

    df = pd.DataFrame(rows)
    # normalize column order
    cols = ["Team Name", "User Domain", "User Name", "User Grade"]
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = df[cols]
    # coerce grades to numeric when possible
    df["User Grade"] = pd.to_numeric(df["User Grade"], errors="coerce")

    # Normalize domains to uppercase (preserve NaN/None)
    def _upper_domain(v):
        if v is None:
            return None
        try:
            return str(v).upper()
        except Exception:
            return v

    df["User Domain"] = df["User Domain"].apply(_upper_domain)

    # --- Compute statistics ---
    report = {}

    # Top 10 grades
    top10 = df.sort_values("User Grade", ascending=False).head(10)
    report["top10"] = top10.to_dict(orient="records")

    # Bottom 10 grades
    bottom10 = df.sort_values("User Grade", ascending=True).head(10)
    report["bottom10"] = bottom10.to_dict(orient="records")

    # Top 12 teams by average team grade
    team_avg = (
        df.groupby("Team Name", dropna=False)["User Grade"].mean().reset_index()
    )
    team_avg = team_avg.rename(columns={"User Grade": "Average Grade"})
    team_avg_sorted = team_avg.sort_values("Average Grade", ascending=False)
    report["top12_teams"] = team_avg_sorted.head(12).to_dict(orient="records")

    # Per-domain top3 and bottom3 by team averages within that domain
    # First, we need team-domain averages: a team may have multiple users in different domains; we compute per-team-per-domain averages
    df_domain = df.dropna(subset=["User Domain"]).copy()
    team_domain_avg = (
        df_domain.groupby(["User Domain", "Team Name"])["User Grade"].mean().reset_index()
    )
    team_domain_avg = team_domain_avg.rename(columns={"User Grade": "Average Grade"})

    per_domain_top3 = {}
    per_domain_bottom3 = {}
    for domain, group in team_domain_avg.groupby("User Domain"):
        sorted_group = group.sort_values("Average Grade", ascending=False)
        per_domain_top3[domain] = sorted_group.head(3).to_dict(orient="records")
        per_domain_bottom3[domain] = sorted_group.tail(3).sort_values("Average Grade", ascending=True).to_dict(orient="records")

    report["per_domain_top3"] = per_domain_top3
    report["per_domain_bottom3"] = per_domain_bottom3

    # --- Statistics: overall and per-domain (min, max, mean, median) ---
    stats = {}
    grades = df["User Grade"].dropna()
    if not grades.empty:
        stats["overall"] = {
            "min": float(grades.min()),
            "max": float(grades.max()),
            "mean": float(grades.mean()),
            "median": float(grades.median()),
            "count": int(grades.count()),
        }
    else:
        stats["overall"] = {"min": None, "max": None, "mean": None, "median": None, "count": 0}

    domain_stats = {}
    if not df_domain.empty:
        grp = df_domain.groupby("User Domain")["User Grade"]
        agg = grp.agg(["min", "max", "mean", "median", "count"]).reset_index()
        # convert numpy types to python native for json
        for _, row in agg.iterrows():
            domain_stats[row["User Domain"]] = {
                "min": float(row["min"]) if not pd.isna(row["min"]) else None,
                "max": float(row["max"]) if not pd.isna(row["max"]) else None,
                "mean": float(row["mean"]) if not pd.isna(row["mean"]) else None,
                "median": float(row["median"]) if not pd.isna(row["median"]) else None,
                "count": int(row["count"]) if not pd.isna(row["count"]) else 0,
            }
    stats["per_domain"] = domain_stats
    report["statistics"] = stats

    # Write Excel with multiple sheets
    out_report_xlsx = root / "summary_report.xlsx"
    with pd.ExcelWriter(out_report_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Grades", index=False)
        # Also write a sheet sorted by domain then user name
        sorted_df = df.sort_values(["User Domain", "User Name"], ascending=[True, True])
        sorted_df.to_excel(writer, sheet_name="All Grades By Domain", index=False)
        top10.to_excel(writer, sheet_name="Top 10 Grades", index=False)
        bottom10.to_excel(writer, sheet_name="Bottom 10 Grades", index=False)
        team_avg_sorted.to_excel(writer, sheet_name="Team Averages", index=False)
        # per-domain sheets
        for domain, group in team_domain_avg.groupby("User Domain"):
            group.sort_values("Average Grade", ascending=False).to_excel(
                writer, sheet_name=f"Domain-{domain}", index=False
            )
        # Statistics sheet: overall and per-domain
        # overall
        overall_df = pd.DataFrame([stats["overall"]])
        overall_df.insert(0, "Scope", "Overall")
        # per-domain
        per_domain_rows = []
        for domain, s in stats["per_domain"].items():
            r = s.copy()
            r["Scope"] = domain
            per_domain_rows.append(r)
        if per_domain_rows:
            per_domain_df = pd.DataFrame(per_domain_rows)
            # reorder columns
            cols_order = ["Scope", "min", "max", "mean", "median", "count"]
            per_domain_df = per_domain_df[cols_order]
            stats_df = pd.concat([overall_df, per_domain_df], ignore_index=True)
        else:
            stats_df = overall_df
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)

    print(f"Wrote {out_report_xlsx}")
    print("Top 10 grades:")
    print(top10.to_string(index=False))


if __name__ == "__main__":
    main(sys.argv[1:])
